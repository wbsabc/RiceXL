import os
import xlrd
import xlwt
import tempfile
from xlwt.Style import XFStyle
from xlutils.copy import copy
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils import column_index_from_string
from RiceXL.RiceFont import RiceFont
from RiceXL.RiceBorder import RiceBorder
from RiceXL.RicePattern import RicePattern
from RiceXL.RiceAlignment import RiceAlignment
from RiceXL.xlutils_newcopy import new_copy
from RiceXL.RiceUtils import get_pos_x_index
from RiceXL.RiceUtils import get_pos_y_index
from RiceXL.RiceUtils import get_right_pos
from RiceXL.RiceUtils import get_down_pos
from RiceXL.RiceUtils import get_corner_positions
from RiceXL.RiceUtils import get_positions_in_area

class RiceExcel(object):
    def __init__(self, file_path, writable=False):
        self.excel_path = file_path
        self.writable = writable
        self.tempfd = None
        self.tempfilename = None
        if self.excel_path.lower().endswith('.xls'):
            self.excel_version = 'xls'
            if os.path.exists(self.excel_path):
                # If formatting_info=False: cannot read styles
                self.workbook = xlrd.open_workbook(self.excel_path, formatting_info=True)
                if writable:
                    # A new workbook for writing
                    self.workbook_write, self.style_list = new_copy(self.workbook)
            elif writable:
                self.workbook_write = xlwt.Workbook(encoding='utf-8')
            else:
                raise Exception('File not found: {}'.format(self.excel_path))   
        elif self.excel_path.lower().endswith('.xlsx'):
            self.excel_version = 'xlsx'
            if os.path.exists(self.excel_path):
                self.workbook = load_workbook(filename=self.excel_path, data_only=True)
            elif writable:
                self.workbook = Workbook()
            else:
                raise Exception('File not found: {}'.format(self.excel_path))
        else:
            raise Exception('Unsupported file format!')
    
    def set_sheet_by_index(self, sheet_index):
        self.sheet_index = sheet_index
        self.sheet_name = self.get_sheet_name_by_index(self.sheet_index)
        if self.excel_version == 'xls':
            self.sheet = self.workbook.sheet_by_index(self.sheet_index)
            if self.writable:
                self.sheet_write = self.workbook_write.get_sheet(self.sheet_index)
        elif self.excel_version == 'xlsx':
            self.sheet = self.workbook.worksheets[self.sheet_index]

    def set_sheet_by_name(self, sheet_name):
        self.sheet_name = sheet_name
        self.sheet_index = self.get_sheet_index_by_name(self.sheet_name)
        if self.excel_version == 'xls':
            self.sheet = self.workbook.sheet_by_name(self.sheet_name)
            if self.writable:
                self.sheet_write = self.workbook_write.get_sheet(self.sheet_name)
        elif self.excel_version == 'xlsx':
            self.sheet = self.workbook[self.sheet_name]
    
    def get_sheet_index_by_name(self, sheet_name):
        if self.excel_version == 'xls':
            sheet_index = self.workbook._sheet_names.index(sheet_name)
        elif self.excel_version == 'xlsx':
            sheet_index = self.workbook.sheetnames.index(sheet_name)
        return sheet_index
    
    def get_sheet_name_by_index(self, sheet_index):
        if self.excel_version == 'xls':
            sheet_name = self.workbook._sheet_names[sheet_index]
        elif self.excel_version == 'xlsx':
            sheet_name = self.workbook.sheetnames[sheet_index]
        return sheet_name
    
    def get_sheet(self, writable=False, sheet_name=None, sheet_index=None):
        if sheet_name is not None and sheet_index is not None:
            raise Exception('Cannot set sheet_index & sheet_name at the same time!')
        if sheet_name is not None:
            if self.excel_version == 'xls':
                if self.writable:
                    return self.workbook_write.get_sheet(sheet_name)
                else:
                    return self.workbook.sheet_by_name(sheet_name)
            elif self.excel_version == 'xlsx':
                return self.workbook[sheet_name]
        elif sheet_index is not None:
            if self.excel_version == 'xls':
                if self.writable:
                    return self.workbook_write.get_sheet(sheet_index)
                else:
                    return self.workbook.sheet_by_index(sheet_index)
            elif self.excel_version == 'xlsx':
                return self.workbook.worksheets[sheet_index]
        else:
            if self.excel_version == 'xls':
                if writable:
                    return self.sheet_write
                else:
                    return self.sheet
            elif self.excel_version == 'xlsx':
                return self.sheet
    
    def get_value(self, position, sheet_index=None, sheet_name=None):
        self.set_sheet_by_index_or_name(sheet_index, sheet_name)
        if self.excel_version == 'xls':
            return self.get_value_xls(position)
        elif self.excel_version == 'xlsx':
            return self.get_value_xlsx(position)
    
    def get_value_xls(self, position):
        pos_x = get_pos_x_index(position)
        pos_y = get_pos_y_index(position)
        pos_x -= 1
        pos_y = int(pos_y) - 1
        
        if pos_y >= self.sheet.nrows or pos_x >= self.sheet.ncols:
            return None
        return self.sheet.cell(pos_y, pos_x).value

    def get_value_xlsx(self, position):
        return self.sheet[position].value

    def set_sheet_by_index_or_name(self, sheet_index, sheet_name):
        if sheet_index is not None and sheet_name is not None:
            raise Exception('Cannot set sheet_index & sheet_name at the same time!')
        if sheet_index is None and sheet_name is None:
            self.check_sheet_set()
        if sheet_index is not None:
            self.set_sheet_by_index(sheet_index)
        if sheet_name is not None:
            self.set_sheet_by_name(sheet_name)
    
    def check_sheet_set(self):
        if self.sheet_index is None or self.sheet_name is None:
            raise Exception('Have not set sheet!')

    def create_sheet(self, sheet_name, use_sheet=False):
        self.check_writable()
        if self.excel_version == 'xls':
            self.sheet_write = self.workbook_write.add_sheet(sheet_name, cell_overwrite_ok=False)
            self.handle_temp_file()
        elif self.excel_version == 'xlsx':
            self.sheet = self.workbook.create_sheet(sheet_name)
        if use_sheet == True:
            self.set_sheet_by_name(sheet_name)
    
    def remove_sheet(self, sheet_name):
        self.check_writable()
        if self.excel_version == 'xls':
            sheet_del = self.get_sheet(True, sheet_name)
            self.workbook_write._Workbook__worksheets.remove(sheet_del)
            self.handle_temp_file()
        elif self.excel_version == 'xlsx':
            sheet_del = self.get_sheet(True, sheet_name)
            self.sheet = self.workbook.remove(sheet_del)
        if sheet_name == self.sheet_name:
            self.sheet_name = None
            self.sheet_index = None
            self.sheet = None
            self.sheet_write = None

    def get_active_sheet(self):
        if self.excel_version == 'xls':
            if self.writable:
                return self.get_sheet(writable=True, sheet_index=self.workbook_write.active_sheet)
            else:
                workbook_write = copy(self.workbook)
                return workbook_write.get_sheet(workbook_write.active_sheet)
        elif self.excel_version == 'xlsx':
            return self.workbook.active
    
    def get_active_sheet_name(self):
        if self.excel_version == 'xls':
            if self.writable:
                return self.get_sheet_name_by_index(self.workbook_write.active_sheet)
            else:
                workbook_write = copy(self.workbook)
                return self.get_sheet_name_by_index(workbook_write.active_sheet)
        elif self.excel_version == 'xlsx':
            return self.get_sheet_name_by_index(self.workbook._active_sheet_index)

    def set_active_sheet(self, sheet_name):
        self.check_writable()
        if self.excel_version == 'xls':
            self.workbook_write.active_sheet = self.get_sheet_index_by_name(sheet_name)
        elif self.excel_version == 'xlsx':
            self.workbook._active_sheet_index = self.get_sheet_index_by_name(sheet_name)
    
    def set_value(self, position, value, style=None):
        self.check_writable()
        if self.excel_version == 'xls':
            pos_x = get_pos_x_index(position) - 1
            pos_y = get_pos_y_index(position) - 1
            if style is None:
                style = self.get_origin_style_by_index(pos_y, pos_x)
                self.sheet_write.write(pos_y, pos_x, value, style)
            elif style == '':
                self.sheet_write.write(pos_y, pos_x, value)
            else:
                self.sheet_write.write(pos_y, pos_x, value, style)
            self.handle_temp_file()
        elif self.excel_version == 'xlsx':
            self.sheet[position].value = value
    
    def handle_temp_file(self):
        if self.tempfd is None:
            self.tempfd, self.tempfilename = tempfile.mkstemp(suffix='.xls', prefix='RiceXL')
        self.save(self.tempfilename, True)
        self.workbook = xlrd.open_workbook(self.tempfilename, formatting_info=True)
        temp_workbook_write, self.style_list = new_copy(self.workbook)
        self.set_sheet_by_index(self.sheet_index)
    
    def get_origin_style_by_position(self, position):
        pos_x = get_pos_x_index(position) - 1
        pos_y = get_pos_y_index(position) - 1        
        return self.get_origin_style_by_index(pos_y, pos_x)

    def get_origin_style_by_index(self, pos_y, pos_x):
        try:
            xf_index = self.sheet.cell_xf_index(pos_y, pos_x)
            style = self.style_list[xf_index]
        except Exception:
            style = XFStyle()
        return style
    
    def set_number_format(self, position, format='General'):
        self.check_writable()
        if self.excel_version == 'xls':
            style = self.get_origin_style_by_position(position)
            value = self.get_value_xls(position)
            style.num_format_str = format
            self.set_value(position, value, style)
        elif self.excel_version == 'xlsx':
            self.sheet[position].number_format = format
    
    def get_number_format(self, position):
        if self.excel_version == 'xls':
            style = self.get_origin_style_by_position(position)
            number_format = style.num_format_str
        elif self.excel_version == 'xlsx':
            number_format = self.sheet[position].number_format
        return number_format
    
    def set_font(self, position, rice_font):
        self.check_writable()
        if self.excel_version == 'xls':
            style = self.get_origin_style_by_position(position)
            value = self.get_value_xls(position)
            style.font = rice_font.font_xls
            self.set_value(position, value, style)
        elif self.excel_version == 'xlsx':
            self.sheet[position].font = rice_font.font_xlsx
    
    def get_font(self, position):
        if self.excel_version == 'xls':
            style = self.get_origin_style_by_position(position)
            rice_font = RiceFont().init_xls(style.font)
        elif self.excel_version == 'xlsx':
            rice_font = RiceFont().init_xlsx(self.sheet[position].font)
        return rice_font
    
    def set_border(self, position, rice_border):
        self.check_writable()
        if self.excel_version == 'xls':
            style = self.get_origin_style_by_position(position)
            value = self.get_value_xls(position)
            style.borders = rice_border.border_xls
            self.set_value(position, value, style)
        elif self.excel_version == 'xlsx':
            self.sheet[position].border = rice_border.border_xlsx
    
    def get_border(self, position):
        if self.excel_version == 'xls':
            style = self.get_origin_style_by_position(position)
            rice_border = RiceBorder().init_xls(style.borders)
        elif self.excel_version == 'xlsx':
            rice_border = RiceBorder().init_xlsx(self.sheet[position].border)
        return rice_border
    
    def set_alignment(self, position, rice_alignment):
        self.check_writable()
        if self.excel_version == 'xls':
            style = self.get_origin_style_by_position(position)
            value = self.get_value_xls(position)
            style.alignment = rice_alignment.alignment_xls
            self.set_value(position, value, style)
        elif self.excel_version == 'xlsx':
            self.sheet[position].alignment = rice_alignment.alignment_xlsx
    
    def get_alignment(self, position):
        if self.excel_version == 'xls':
            style = self.get_origin_style_by_position(position)
            rice_alignment = RiceAlignment().init_xls(style.alignment)
        elif self.excel_version == 'xlsx':
            rice_alignment = RiceAlignment().init_xlsx(self.sheet[position].alignment)
        return rice_alignment
    
    def set_pattern(self, position, rice_pattern):
        self.check_writable()
        if self.excel_version == 'xls':
            style = self.get_origin_style_by_position(position)
            value = self.get_value_xls(position)
            style.pattern = rice_pattern.pattern_xls
            self.set_value(position, value, style)
        elif self.excel_version == 'xlsx':
            self.sheet[position].fill = rice_pattern.pattern_xlsx
    
    def get_pattern(self, position):
        if self.excel_version == 'xls':
            style = self.get_origin_style_by_position(position)
            rice_pattern = RicePattern().init_xls(style.pattern)
        elif self.excel_version == 'xlsx':
            rice_pattern = RicePattern().init_xlsx(self.sheet[position].fill)
        return rice_pattern
    
    def set_height(self, row, height):
        self.check_writable()
        if self.excel_version == 'xls':
            row -= 1
            self.sheet_write.row(row).height_mismatch = True
            self.sheet_write.row(row).height = int(20 * height)
            self.handle_temp_file()
        elif self.excel_version == 'xlsx':
            self.sheet.row_dimensions[row].height = height
    
    def get_height(self, row):
        if self.excel_version == 'xls':
            row -= 1
            if row >= len(self.sheet.rowinfo_map):
                height = 20 * 12.75
            else:
                height = self.sheet.rowinfo_map[row].height
            height /= 20
        elif self.excel_version == 'xlsx':
            height = self.sheet.row_dimensions[row].height
            if height is None:
                height = 13.5
        return height
    
    def set_width(self, col, width):
        self.check_writable()
        if self.excel_version == 'xls':
            col_index = column_index_from_string(col) - 1
            self.sheet_write.col(col_index).width = int(256 * width)
            self.handle_temp_file()
        elif self.excel_version == 'xlsx':
            self.sheet.column_dimensions[col].width = width
    
    def get_width(self, col):
        if self.excel_version == 'xls':
            col_index = column_index_from_string(col) - 1
            if self.sheet.colinfo_map.get(col_index) is not None:
                width = round(self.sheet.colinfo_map[col_index].width / 256, 1)
            else:
                width = 9.1
        elif self.excel_version == 'xlsx':
            width = self.sheet.column_dimensions[col].width
        return width
    
    def hide_row(self, row):
        self.check_writable()
        if self.excel_version == 'xls':
            self.set_height(row, 0)
        elif self.excel_version == 'xlsx':
            self.sheet.row_dimensions[row].hidden = True
    
    def unhide_row(self, row):
        self.check_writable()
        if self.excel_version == 'xls':
            height = self.get_height(row)
            if height == 0:
                height = 12.75
            self.set_height(row, height)
        elif self.excel_version == 'xlsx':
            self.sheet.row_dimensions[row].hidden = False
    
    def is_row_hidden(self, row):
        if self.excel_version == 'xls':
            height = self.get_height(row)
            if height == 0:
                hidden = True
            else:
                hidden = False
        elif self.excel_version == 'xlsx':
            hidden = self.sheet.row_dimensions[row].hidden
        return hidden
    
    def hide_col(self, col):
        self.check_writable()
        if self.excel_version == 'xls':
            self.set_width(col, 0)
        elif self.excel_version == 'xlsx':
            self.sheet.column_dimensions[col].hidden = True
    
    def unhide_col(self, col):
        self.check_writable()
        if self.excel_version == 'xls':
            width = self.get_width(col)
            if width == 0:
                width = 9.1
            self.set_width(col, width)
        elif self.excel_version == 'xlsx':
            self.sheet.column_dimensions[col].hidden = False
    
    def is_col_hidden(self, col):
        if self.excel_version == 'xls':
            width = self.get_width(col)
            if width == 0:
                hidden = True
            else:
                hidden = False
        elif self.excel_version == 'xlsx':
            hidden = self.sheet.column_dimensions[col].hidden
        return hidden
    
    def merge_cells(self, pos_begin, pos_end):
        begin_x_index, end_x_index, begin_y_index, end_y_index = get_corner_positions(pos_begin, pos_end)
        if self.excel_version == 'xls':
            pos_first = get_column_letter(begin_x_index) + str(begin_y_index)
            value = self.get_value_xls(pos_first)
            style = self.get_origin_style_by_position(pos_first)
            begin_x_index -= 1
            end_x_index -= 1
            begin_y_index -= 1
            end_y_index -= 1
            self.sheet_write.write_merge(begin_y_index, end_y_index, begin_x_index, end_x_index, value, style)
            self.handle_temp_file()
        elif self.excel_version == 'xlsx':
            self.sheet.merge_cells(start_row=begin_y_index, start_column=begin_x_index, end_row=end_y_index, end_column=end_x_index)
    
    def unmerge_cells(self, pos_begin, pos_end):
        begin_x_index, end_x_index, begin_y_index, end_y_index = get_corner_positions(pos_begin, pos_end)
        if self.excel_version == 'xls':
            begin_x_index -= 1
            end_x_index -= 1
            begin_y_index -= 1
            end_y_index -= 1
            self.sheet_write.merged_ranges.remove((begin_y_index, end_y_index, begin_x_index, end_x_index))
            self.handle_temp_file()
        elif self.excel_version == 'xlsx':
            self.sheet.unmerge_cells(start_row=begin_y_index, start_column=begin_x_index, end_row=end_y_index, end_column=end_x_index)

    def save(self, new_file_path=None, temp_flag=False):
        self.check_writable()
        if new_file_path is None:
            new_file_path = self.excel_path
        if self.excel_version == 'xls':
            self.workbook_write.save(new_file_path)
            if not temp_flag:
                self.workbook_write = copy(self.workbook)
        elif self.excel_version == 'xlsx':
            self.workbook.save(new_file_path)

    def clean_up(self):
        if self.tempfd is not None:
            os.close(self.tempfd)
            os.unlink(self.tempfilename)
            self.tempfd = None
            self.tempfilename = None

    def check_writable(self):
        if not self.writable:
            raise Exception('The file is not writable!')
    
    def sum(self, *positions):
        self.check_sheet_set()
        sum = 0
        for pos in positions:
            sum += float(self.get_value(pos))
        return sum
    
    def sum_in_area(self, pos_begin, pos_end):
        self.check_sheet_set()
        pos_list = get_positions_in_area(pos_begin, pos_end)
        sum = 0
        for pos in pos_list:
            sum += float(self.get_value(pos))
        return sum
    
    def get_values_matrix_in_area(self, pos_begin, pos_end):
        self.check_sheet_set()
        begin_x_index, end_x_index, begin_y_index, end_y_index = get_corner_positions(pos_begin, pos_end)
        value_matrix = []
        for x in range(begin_x_index, end_x_index + 1):
            value_list = []
            for y in range(begin_y_index, end_y_index + 1):
                pos = get_column_letter(x) + str(y)
                value = self.get_value(pos)
                value_list.append(value)
            value_matrix.append(value_list)
        return value_matrix
    
    def set_value_by_area(self, pos_begin, value_matrix):
        self.check_sheet_set()
        pos_cursor = pos_begin
        for value_list in value_matrix:
            pos = pos_cursor
            for value in value_list:
                self.set_value(pos, value)
                pos = get_down_pos(pos)
            pos_cursor = get_right_pos(pos_cursor)