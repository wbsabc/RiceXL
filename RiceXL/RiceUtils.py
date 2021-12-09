import re
from openpyxl.styles.colors import COLOR_INDEX
from openpyxl.utils import get_column_letter
from openpyxl.utils import column_index_from_string

def get_color_rgb(color):
    # Only support rgb
    if color.type == 'rgb':
        color = '00' + color.rgb[2:]
    else:
        color = '00000000'
    return color

def get_color_index_by_rgb(color):
    return COLOR_INDEX.index(color)

def get_color_rgb_by_index(index):
    if index == 32767:
        index = 0
    return COLOR_INDEX[index % 64]

def get_pos_x(position):
    return re.findall(r'\D+', position)[0] # Match letters

# Start with 1
def get_pos_x_index(position):
    return column_index_from_string(get_pos_x(position))

# Start with 1
def get_pos_y(position):
    return re.findall(r'\d+', position)[0] # Match numbers

def get_pos_y_index(position):
    return int(get_pos_y(position))

def get_left_pos(position, shift=1):
    shift = -shift
    return get_right_pos(position, shift)

def get_right_pos(position, shift=1):
    pos_x = get_pos_x_index(position)
    pos_y = get_pos_y(position)

    pos_x = get_column_letter(pos_x + shift)
    return pos_x + pos_y

def get_up_pos(position, shift=1):
    shift = -shift
    return get_down_pos(position, shift)

def get_down_pos(position, shift=1):
    pos_x = get_pos_x(position)
    pos_y = get_pos_y_index(position)
    pos_y += shift
    if pos_y <= 0:
        raise Exception('Invalid row index %d' % pos_y)

    return pos_x + str(pos_y)

def get_corner_positions(pos_begin, pos_end):
    pos_begin_x_index = get_pos_x_index(pos_begin)
    pos_begin_y_index = get_pos_y_index(pos_begin)
    pos_end_x_index = get_pos_x_index(pos_end)
    pos_end_y_index = get_pos_y_index(pos_end)
    if pos_begin_x_index <= pos_end_x_index:
        begin_x_index = pos_begin_x_index
        end_x_index = pos_end_x_index
    else:
        begin_x_index = pos_end_x_index
        end_x_index = pos_begin_x_index
    if pos_begin_y_index <= pos_end_y_index:
        begin_y_index = pos_begin_y_index
        end_y_index = pos_end_y_index
    else:
        begin_y_index = pos_end_y_index
        end_y_index = pos_begin_y_index
    return begin_x_index, end_x_index, begin_y_index, end_y_index

def get_positions_in_area(pos_begin, pos_end):
    begin_x_index, end_x_index, begin_y_index, end_y_index = get_corner_positions(pos_begin, pos_end)
    pos_list = []
    for x in range(begin_x_index, end_x_index + 1):
        for y in range(begin_y_index, end_y_index + 1):
            pos = get_column_letter(x) + str(y)
            pos_list.append(pos)
    return pos_list

def get_positions_matrix_in_area(pos_begin, pos_end):
    begin_x_index, end_x_index, begin_y_index, end_y_index = get_corner_positions(pos_begin, pos_end)
    pos_matrix = []
    for x in range(begin_x_index, end_x_index + 1):
        pos_list = []
        for y in range(begin_y_index, end_y_index + 1):
            pos = get_column_letter(x) + str(y)
            pos_list.append(get_column_letter(x) + str(y))
        pos_matrix.append(pos_list)
    return pos_matrix

def none_to_zero(value):
    if value is None:
        return 0
    else:
        return value